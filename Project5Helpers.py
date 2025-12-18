#this is a set of helper functions for Project 5
#you don't need to understand this code, we haven't covered everything in class
import os.path
import openpyxl

def get_job_data()->list[dict]:
    initial_data = get_salary_data_from_file(os.path.join("Data", "MA_TECH_Salaries.xlsx"))
    for job_info in initial_data:
        match job_info["code"]:
            case "15-1212":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_InfoSecEng.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1244" | "15-1241":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_PenTesters.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1299":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_DigitalForensics.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1252":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_SoftwareDev.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1242"|"15-1243":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_DBA.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1253":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_SDET.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1254":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_WebDev.xlsx"))
                job_info["needed_skills"] = skills_data
            case "15-1211":
                skills_data = get_skills_list(os.path.join("Data", "tech_skills_SysAn.xlsx"))
                job_info["needed_skills"] = skills_data
    return initial_data


def get_salary_data_from_file(filename:str)->list[dict]:
    data = []
    salary_excel = openpyxl.load_workbook(filename)
    datasource = salary_excel.active
    for row in datasource.iter_rows(min_row=2):
        if row[0].value == None:
            break
        job_info = {
            "state":row[1].value,
            "code":row[2].value,
            "job_class":row[3].value,
            "ma_emp_num":row[5].value,
            "ave_annual_salary":row[6].value,
            "median_annual_salary":row[9].value,
            "starting_salary":row[7].value,
            "top_salary":row[11].value
        }
        if job_info["top_salary"] == '#': #there is one bad data - fix it
            job_info["top_salary"] = 220_000
        data.append(job_info)
    return data

def get_skills_list(file:str)->list[str]:
    skills = []
    skill_excel = openpyxl.load_workbook(file)
    datasource = skill_excel.active
    for row in datasource.iter_rows(min_row=5):
        this_skill = {
            "category":row[1].value,
            "skill":row[2].value
        }
        skills.append(this_skill)
    return skills

